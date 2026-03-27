from flask import (Flask, render_template, request, redirect,
                   session, url_for, flash, send_file)
import sqlite3, os, io, json
from functools import wraps
from datetime import datetime, date
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'BCA_ATTENDANCE_SYSTEM_2026_SECURE_KEY')

# Always resolve DB path relative to this file, not cwd
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH  = os.path.join(BASE_DIR, 'data', 'attendance.db')

# ─────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────
def get_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


# Period timing definitions
FN_PERIODS = {
    1: ('09:15', '10:05'),
    2: ('10:05', '10:55'),
    # Break 10:55 – 11:10
    3: ('11:10', '12:00'),
    4: ('12:00', '12:50'),
}
AF_PERIODS = {
    1: ('13:30', '14:20'),
    2: ('14:20', '15:10'),
    # Break 15:10 – 15:25
    3: ('15:25', '16:15'),
    4: ('16:15', '17:05'),
}

def get_current_period(shift):
    """Return current period number (1-4) for the given shift, or None if outside/break."""
    now = datetime.now()
    now_t = now.strftime('%H:%M')
    periods = FN_PERIODS if shift == 'FN' else AF_PERIODS
    for period_no, (start, end) in periods.items():
        if start <= now_t < end:
            return period_no
    return None


def get_period_label(shift, period_no):
    periods = FN_PERIODS if shift == 'FN' else AF_PERIODS
    t = periods.get(period_no)
    if t:
        return f"Period {period_no}  ({t[0]} – {t[1]})"
    return f"Period {period_no}"


def login_required(role=None):
    """Decorator factory for role-based protection."""
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if 'user_id' not in session:
                flash('Please log in first.', 'warning')
                return redirect('/')
            if role and session.get('role') != role:
                flash('Access denied.', 'danger')
                return redirect('/')
            return f(*args, **kwargs)
        return wrapped
    return decorator


def calc_percentage(present, total):
    """total here means countable classes (P + A only, not OD/L)."""
    if total == 0:
        return 0.0
    return round((present / total) * 100, 1)

def classes_needed_for_75(present, total):
    """How many consecutive classes a student needs to attend to hit 75%."""
    if total == 0:
        return 0
    if calc_percentage(present, total) >= 75:
        return 0
    needed = 0
    while True:
        needed += 1
        if calc_percentage(present + needed, total + needed) >= 75:
            return needed


def status_color(pct):
    if pct >= 75:
        return 'green'
    elif pct >= 50:
        return 'yellow'
    return 'red'


def init_admin():
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE role='admin'").fetchone()
    if not existing:
        db.execute(
            "INSERT INTO users (name,username,password,role,status) VALUES (?,?,?,'admin','active')",
            ('System Admin', 'ADMIN', generate_password_hash('Admin12345'))
        )
        db.commit()
    db.close()


# ─────────────────────────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────────────────────────
@app.route('/')
def index():
    if 'user_id' in session:
        return redirect(url_for(session['role'] + '_dashboard'))
    return render_template('login.html')


@app.route('/login', methods=['POST'])
def login():
    username = request.form['username'].strip().upper()
    password = request.form['password']
    db = get_db()
    user = db.execute('SELECT * FROM users WHERE UPPER(username)=?', (username,)).fetchone()
    db.close()
    if user and user['status'] == 'inactive':
        flash('Your account is deactivated. Contact Admin.', 'danger')
        return redirect('/')
    if user and check_password_hash(user['password'], password):
        session.update({
            'user_id': user['id'],
            'role': user['role'],
            'name': user['name'],
            'username': user['username']
        })
        return redirect(url_for(user['role'] + '_dashboard'))
    flash('Invalid username or password.', 'danger')
    return redirect('/')


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/')


# ─────────────────────────────────────────────────────────────────
# ADMIN — DASHBOARD
# ─────────────────────────────────────────────────────────────────
@app.route('/admin_dashboard')
@login_required('admin')
def admin_dashboard():
    db = get_db()
    total_students  = db.execute("SELECT COUNT(*) FROM students").fetchone()[0]
    total_teachers  = db.execute("SELECT COUNT(*) FROM users WHERE role='teacher'").fetchone()[0]
    total_subjects  = db.execute("SELECT COUNT(*) FROM subjects").fetchone()[0]
    today           = date.today().isoformat()
    today_records   = db.execute("SELECT COUNT(*) FROM attendance WHERE date=?", (today,)).fetchone()[0]
    today_present   = db.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='P'", (today,)).fetchone()[0]
    today_absent    = db.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='A'", (today,)).fetchone()[0]
    today_od        = db.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='OD'", (today,)).fetchone()[0]
    today_leave     = db.execute("SELECT COUNT(*) FROM attendance WHERE date=? AND status='L'", (today,)).fetchone()[0]

    # Weekly attendance trend (last 7 days) — ascending for chart display
    weekly_raw = db.execute("""
        SELECT date,
               SUM(CASE WHEN status='P'  THEN 1 ELSE 0 END) as p,
               SUM(CASE WHEN status='A'  THEN 1 ELSE 0 END) as a,
               SUM(CASE WHEN status='OD' THEN 1 ELSE 0 END) as od,
               SUM(CASE WHEN status='L'  THEN 1 ELSE 0 END) as lv
        FROM attendance
        GROUP BY date ORDER BY date DESC LIMIT 7
    """).fetchall()
    weekly = list(reversed(weekly_raw))  # oldest → newest for chart

    # For chart.js — serialize to JSON-safe lists
    weekly_labels  = json.dumps([w['date'] for w in weekly])
    weekly_present = json.dumps([w['p']    for w in weekly])
    weekly_absent  = json.dumps([w['a']    for w in weekly])
    weekly_od      = json.dumps([w['od']   for w in weekly])
    weekly_leave   = json.dumps([w['lv']   for w in weekly])

    db.close()
    return render_template('admin_dashboard.html',
                           total_students=total_students,
                           total_teachers=total_teachers,
                           total_subjects=total_subjects,
                           today_records=today_records,
                           today_present=today_present,
                           today_absent=today_absent,
                           today_od=today_od,
                           today_leave=today_leave,
                           weekly=weekly_raw,
                           weekly_labels=weekly_labels,
                           weekly_present=weekly_present,
                           weekly_absent=weekly_absent,
                           weekly_od=weekly_od,
                           weekly_leave=weekly_leave,
                           now=datetime.now())


# ─────────────────────────────────────────────────────────────────
# ADMIN — STUDENTS
# ─────────────────────────────────────────────────────────────────
@app.route('/manage_students')
@login_required('admin')
def manage_students():
    q   = request.args.get('q', '').strip()
    db  = get_db()
    if q:
        students = db.execute(
            "SELECT * FROM students WHERE UPPER(name) LIKE ? OR UPPER(reg_no) LIKE ? ORDER BY year,section,reg_no",
            (f'%{q.upper()}%', f'%{q.upper()}%')
        ).fetchall()
    else:
        students = db.execute("SELECT * FROM students ORDER BY year,section,reg_no").fetchall()
    db.close()
    return render_template('manage_students.html', students=students, q=q)


@app.route('/add_student', methods=['POST'])
@login_required('admin')
def add_student():
    reg_no  = request.form['reg_no'].strip().upper()
    name    = request.form['name'].strip()
    year    = int(request.form['year'])
    semester= int(request.form['semester'])
    section = request.form['section'].strip().upper()
    shift   = request.form['shift']
    contact = request.form.get('contact', '').strip()
    email   = request.form.get('email', '').strip()
    db = get_db()
    try:
        db.execute(
            "INSERT INTO students VALUES (?,?,?,?,?,?,?,?)",
            (reg_no, name, year, semester, section, shift, contact, email)
        )
        # Create user account for student
        db.execute(
            "INSERT INTO users (name,username,password,role,status) VALUES (?,?,?,'student','active')",
            (name, reg_no, generate_password_hash(reg_no[-4:]))
        )
        db.commit()
        flash(f'✅ Student {name} added. Login: {reg_no} | Password: {reg_no[-4:]}', 'success')
    except sqlite3.IntegrityError:
        flash(f'❌ Register number {reg_no} already exists.', 'danger')
    db.close()
    return redirect(url_for('manage_students'))


@app.route('/edit_student/<reg_no>', methods=['GET','POST'])
@login_required('admin')
def edit_student(reg_no):
    db = get_db()
    if request.method == 'POST':
        name    = request.form['name'].strip()
        year    = int(request.form['year'])
        semester= int(request.form['semester'])
        section = request.form['section'].strip().upper()
        shift   = request.form['shift']
        contact = request.form.get('contact','').strip()
        email   = request.form.get('email','').strip()
        db.execute(
            "UPDATE students SET name=?,year=?,semester=?,section=?,shift=?,contact=?,email=? WHERE reg_no=?",
            (name, year, semester, section, shift, contact, email, reg_no)
        )
        db.execute("UPDATE users SET name=? WHERE username=?", (name, reg_no))
        db.commit()
        db.close()
        flash('✅ Student updated.', 'success')
        return redirect(url_for('manage_students'))
    student = db.execute("SELECT * FROM students WHERE reg_no=?", (reg_no,)).fetchone()
    db.close()
    return render_template('edit_student.html', student=student)


@app.route('/delete_student/<reg_no>')
@login_required('admin')
def delete_student(reg_no):
    db = get_db()
    db.execute("DELETE FROM attendance WHERE student_id=?", (reg_no,))
    db.execute("DELETE FROM students WHERE reg_no=?", (reg_no,))
    db.execute("DELETE FROM users WHERE username=? AND role='student'", (reg_no,))
    db.commit()
    db.close()
    flash('🗑 Student deleted.', 'info')
    return redirect(url_for('manage_students'))


# ─────────────────────────────────────────────────────────────────
# ADMIN — TEACHERS
# ─────────────────────────────────────────────────────────────────
@app.route('/manage_teachers')
@login_required('admin')
def manage_teachers():
    db = get_db()
    teachers = db.execute("SELECT * FROM users WHERE role='teacher' ORDER BY id").fetchall()
    db.close()
    return render_template('manage_teachers.html', teachers=teachers)


@app.route('/add_teacher', methods=['POST'])
@login_required('admin')
def add_teacher():
    name = request.form['name'].strip()
    db   = get_db()
    count = db.execute("SELECT COUNT(*) FROM users WHERE role='teacher'").fetchone()[0]
    n    = count + 1
    username = f'TEACHER{n}'
    password = f'Teacher@{n}'
    db.execute(
        "INSERT INTO users (name,username,password,role,status) VALUES (?,?,?,'teacher','active')",
        (name, username, generate_password_hash(password))
    )
    db.commit()
    db.close()
    flash(f'✅ Teacher added! Login → Username: {username}  |  Password: {password}', 'success')
    return redirect(url_for('manage_teachers'))


@app.route('/edit_teacher/<int:tid>', methods=['GET','POST'])
@login_required('admin')
def edit_teacher(tid):
    db = get_db()
    if request.method == 'POST':
        name = request.form['name'].strip()
        db.execute("UPDATE users SET name=? WHERE id=?", (name, tid))
        db.commit()
        db.close()
        flash('✅ Teacher name updated.', 'success')
        return redirect(url_for('manage_teachers'))
    teacher = db.execute("SELECT * FROM users WHERE id=? AND role='teacher'", (tid,)).fetchone()
    db.close()
    return render_template('edit_teacher.html', teacher=teacher)


@app.route('/delete_teacher/<int:tid>')
@login_required('admin')
def delete_teacher(tid):
    db = get_db()
    db.execute("DELETE FROM timetable WHERE teacher_id=?", (tid,))
    db.execute("DELETE FROM users WHERE id=? AND role='teacher'", (tid,))
    db.commit()
    db.close()
    flash('🗑 Teacher deleted.', 'info')
    return redirect(url_for('manage_teachers'))


@app.route('/toggle_status/<int:uid>')
@login_required('admin')
def toggle_status(uid):
    db = get_db()
    user = db.execute("SELECT status FROM users WHERE id=?", (uid,)).fetchone()
    if user:
        new_status = 'inactive' if user['status'] == 'active' else 'active'
        db.execute("UPDATE users SET status=? WHERE id=?", (new_status, uid))
        db.commit()
        flash(f'Account {'activated' if new_status=="active" else "deactivated"}.', 'info')
    db.close()
    return redirect(request.referrer or url_for('manage_teachers'))


# ─────────────────────────────────────────────────────────────────
# ADMIN — SUBJECTS
# ─────────────────────────────────────────────────────────────────
@app.route('/manage_subjects')
@login_required('admin')
def manage_subjects():
    db = get_db()
    subjects = db.execute("SELECT * FROM subjects ORDER BY year,semester,id").fetchall()
    db.close()
    return render_template('manage_subjects.html', subjects=subjects)


@app.route('/add_subject', methods=['POST'])
@login_required('admin')
def add_subject():
    name     = request.form['name'].strip()
    year     = int(request.form['year'])
    semester = int(request.form['semester'])
    db = get_db()
    db.execute("INSERT INTO subjects (name,year,semester) VALUES (?,?,?)", (name, year, semester))
    db.commit()
    db.close()
    flash('✅ Subject added.', 'success')
    return redirect(url_for('manage_subjects'))


@app.route('/delete_subject/<int:sid>')
@login_required('admin')
def delete_subject(sid):
    db = get_db()
    db.execute("DELETE FROM timetable WHERE subject_id=?", (sid,))
    db.execute("DELETE FROM subjects WHERE id=?", (sid,))
    db.commit()
    db.close()
    flash('🗑 Subject deleted.', 'info')
    return redirect(url_for('manage_subjects'))


# ─────────────────────────────────────────────────────────────────
# ADMIN — TIMETABLE
# ─────────────────────────────────────────────────────────────────
@app.route('/manage_timetable')
@login_required('admin')
def manage_timetable():
    db = get_db()
    timetable = db.execute("""
        SELECT tt.*, u.name as tname, s.name as sname
        FROM timetable tt
        JOIN users u ON tt.teacher_id = u.id
        JOIN subjects s ON tt.subject_id = s.id
        ORDER BY
          CASE tt.day WHEN 'Monday' THEN 1 WHEN 'Tuesday' THEN 2 WHEN 'Wednesday' THEN 3
                      WHEN 'Thursday' THEN 4 WHEN 'Friday' THEN 5 ELSE 6 END,
          tt.shift, tt.period
    """).fetchall()
    teachers  = db.execute("SELECT id,name FROM users WHERE role='teacher' AND status='active'").fetchall()
    subjects  = db.execute("SELECT id,name,year,semester FROM subjects ORDER BY year,semester,name").fetchall()
    db.close()
    days = ['Monday','Tuesday','Wednesday','Thursday','Friday']
    return render_template('manage_timetable.html',
                           timetable=timetable, teachers=teachers,
                           subjects=subjects, days=days)


@app.route('/add_timetable', methods=['POST'])
@login_required('admin')
def add_timetable():
    day        = request.form['day']
    period     = int(request.form['period'])
    shift      = request.form['shift']
    year       = int(request.form['year'])
    section    = request.form['section'].strip().upper()
    teacher_id = int(request.form['teacher_id'])
    subject_id = int(request.form['subject_id'])
    db = get_db()
    # Check for conflict
    conflict = db.execute(
        "SELECT id FROM timetable WHERE day=? AND period=? AND shift=? AND year=? AND section=?",
        (day, period, shift, year, section)
    ).fetchone()
    if conflict:
        flash('⚠ A timetable entry already exists for that slot. Delete it first.', 'warning')
    else:
        db.execute(
            "INSERT INTO timetable (day,period,shift,year,section,teacher_id,subject_id) VALUES (?,?,?,?,?,?,?)",
            (day, period, shift, year, section, teacher_id, subject_id)
        )
        db.commit()
        flash('✅ Timetable entry added.', 'success')
    db.close()
    return redirect(url_for('manage_timetable'))


@app.route('/delete_timetable/<int:tid>')
@login_required('admin')
def delete_timetable(tid):
    db = get_db()
    db.execute("DELETE FROM timetable WHERE id=?", (tid,))
    db.commit()
    db.close()
    flash('🗑 Timetable entry removed.', 'info')
    return redirect(url_for('manage_timetable'))


# ─────────────────────────────────────────────────────────────────
# ADMIN — PASSWORD RESET
# ─────────────────────────────────────────────────────────────────
@app.route('/reset_password', methods=['GET','POST'])
@login_required('admin')
def reset_password():
    db = get_db()
    if request.method == 'POST':
        uid      = int(request.form['uid'])
        new_pass = request.form['new_password'].strip()
        if len(new_pass) < 4:
            flash('Password must be at least 4 characters.', 'danger')
        else:
            db.execute("UPDATE users SET password=? WHERE id=?", (generate_password_hash(new_pass), uid))
            db.commit()
            flash('✅ Password reset successfully.', 'success')
    users = db.execute("SELECT id,name,username,role FROM users ORDER BY role,name").fetchall()
    db.close()
    return render_template('reset_password.html', users=users)


# ─────────────────────────────────────────────────────────────────
# ADMIN — REPORTS
# ─────────────────────────────────────────────────────────────────
@app.route('/reports')
@login_required('admin')
def reports():
    db  = get_db()
    year    = request.args.get('year', type=int)
    section = request.args.get('section', '')
    sub_id  = request.args.get('subject_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to   = request.args.get('date_to', '')

    students_q = "SELECT * FROM students"
    params = []
    clauses = []
    if year:
        clauses.append("year=?"); params.append(year)
    if section:
        clauses.append("section=?"); params.append(section.upper())
    if clauses:
        students_q += " WHERE " + " AND ".join(clauses)
    students_q += " ORDER BY reg_no"
    students = db.execute(students_q, params).fetchall()

    subjects = db.execute("SELECT * FROM subjects ORDER BY year,semester,name").fetchall()
    summary  = []
    for s in students:
        att_q = "SELECT status FROM attendance WHERE student_id=?"
        att_p = [s['reg_no']]
        if sub_id:
            att_q += " AND subject_id=?"; att_p.append(sub_id)
        if date_from:
            att_q += " AND date >= ?"; att_p.append(date_from)
        if date_to:
            att_q += " AND date <= ?"; att_p.append(date_to)
        
        records = db.execute(att_q, att_p).fetchall()
        countable = [r for r in records if r['status'] in ('P', 'A')]
        total   = len(countable)
        present = sum(1 for r in countable if r['status'] == 'P')
        od      = sum(1 for r in records if r['status'] == 'OD')
        lv      = sum(1 for r in records if r['status'] == 'L')
        pct     = calc_percentage(present, total)
        summary.append({
            'reg_no': s['reg_no'], 'name': s['name'],
            'year': s['year'], 'section': s['section'],
            'present': present, 'total': total, 'od': od, 'lv': lv, 'pct': pct,
            'color': status_color(pct)
        })

    # Overall chart data matching filters
    att_base_q = "SELECT status FROM attendance WHERE 1=1"
    att_base_p = []
    if date_from:
        att_base_q += " AND date >= ?"; att_base_p.append(date_from)
    if date_to:
        att_base_q += " AND date <= ?"; att_base_p.append(date_to)
    
    all_att   = db.execute(att_base_q, att_base_p).fetchall()
    chart_p   = sum(1 for r in all_att if r['status'] == 'P')
    chart_a   = sum(1 for r in all_att if r['status'] == 'A')
    chart_od  = sum(1 for r in all_att if r['status'] == 'OD')
    chart_lv  = sum(1 for r in all_att if r['status'] == 'L')

    sections = [chr(c) for c in range(ord('A'), ord('M'))]
    db.close()
    return render_template('reports.html',
                           summary=summary, subjects=subjects,
                           chart_p=chart_p, chart_a=chart_a,
                           chart_od=chart_od, chart_lv=chart_lv,
                           sections=sections,
                           sel_year=year, sel_section=section, sel_sub=sub_id,
                           sel_date_from=date_from, sel_date_to=date_to)


# ─────────────────────────────────────────────────────────────────
# EXPORT — EXCEL
# ─────────────────────────────────────────────────────────────────
@app.route('/export/excel')
@login_required('admin')
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        flash('openpyxl not installed. Run: pip install openpyxl', 'danger')
        return redirect(url_for('reports'))

    db   = get_db()
    rows = db.execute("""
        SELECT s.reg_no, s.name, s.year, s.section, s.shift,
               sub.name as subject, a.date, a.period, a.status
        FROM attendance a
        JOIN students s   ON a.student_id = s.reg_no
        JOIN subjects sub ON a.subject_id = sub.id
        ORDER BY s.year, s.section, s.reg_no, a.date, a.period
    """).fetchall()
    db.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    # Header
    headers = ['Reg No','Name','Year','Section','Shift','Subject','Date','Period','Status']
    hdr_fill = PatternFill(fill_type='solid', fgColor='3C4B9B')
    hdr_font = Font(color='FFFFFF', bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = max(15, len(h)+2)

    green_fill = PatternFill(fill_type='solid', fgColor='C8F7C5')
    red_fill   = PatternFill(fill_type='solid', fgColor='FADBD8')
    blue_fill  = PatternFill(fill_type='solid', fgColor='D6EAF8')
    purple_fill= PatternFill(fill_type='solid', fgColor='E8DAEF')

    for r_idx, row in enumerate(rows, 2):
        data = [row['reg_no'], row['name'], row['year'], row['section'],
                row['shift'], row['subject'], row['date'], row['period'], row['status']]
        for c_idx, val in enumerate(data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if c_idx == 9:  # Status column
                if val == 'P':
                    cell.fill = green_fill
                elif val == 'A':
                    cell.fill = red_fill
                elif val == 'OD':
                    cell.fill = blue_fill
                elif val == 'L':
                    cell.fill = purple_fill
                cell.alignment = Alignment(horizontal='center')

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"Attendance_{date.today().isoformat()}.xlsx"
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ─────────────────────────────────────────────────────────────────
# EXPORT — PDF
# ─────────────────────────────────────────────────────────────────
@app.route('/export/pdf')
@login_required('admin')
def export_pdf():
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
    except ImportError:
        flash('reportlab not installed. Run: pip install reportlab', 'danger')
        return redirect(url_for('reports'))

    db = get_db()
    rows = db.execute("""
        SELECT s.reg_no, s.name, s.year, s.section,
               sub.name as subject,
               SUM(CASE WHEN a.status='P' THEN 1 ELSE 0 END) as present,
               SUM(CASE WHEN a.status IN ('P','A') THEN 1 ELSE 0 END) as total,
               SUM(CASE WHEN a.status='OD' THEN 1 ELSE 0 END) as od,
               SUM(CASE WHEN a.status='L' THEN 1 ELSE 0 END) as leave
        FROM students s
        LEFT JOIN attendance a ON s.reg_no = a.student_id
        LEFT JOIN subjects sub ON a.subject_id = sub.id
        GROUP BY s.reg_no, sub.id
        ORDER BY s.year, s.section, s.reg_no
    """).fetchall()
    db.close()

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    story  = []
    story.append(Paragraph('BCA Online Attendance System — Report', styles['Title']))
    story.append(Paragraph(f'Generated: {date.today().strftime("%d %B %Y")}', styles['Normal']))
    story.append(Spacer(1, 20))

    table_data = [['Reg No','Name','Year','Section','Subject','Present','Total','OD','Leave','%']]
    for row in rows:
        total   = row['total'] or 0
        present = row['present'] or 0
        od      = row['od'] or 0
        leave   = row['leave'] or 0
        pct     = calc_percentage(present, total)
        table_data.append([row['reg_no'], row['name'], row['year'], row['section'],
                            row['subject'] or '—', present, total, od, leave, f'{pct}%'])

    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3C4B9B')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,0), 10),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F4F7FA')]),
        ('GRID',       (0,0), (-1,-1), 0.25, colors.lightgrey),
        ('ALIGN',      (2,0), (-1,-1), 'CENTER'),
        ('FONTSIZE',   (0,1), (-1,-1), 9),
    ]))
    story.append(tbl)
    doc.build(story)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f'Report_{date.today().isoformat()}.pdf',
                     mimetype='application/pdf')


# ─────────────────────────────────────────────────────────────────
# TEACHER — DASHBOARD
# ─────────────────────────────────────────────────────────────────
@app.route('/teacher_dashboard')
@login_required('teacher')
def teacher_dashboard():
    db          = get_db()
    teacher_id  = session['user_id']
    today_name  = datetime.now().strftime('%A')   # e.g. 'Monday'
    today_date  = date.today().isoformat()

    # Check both shifts; prefer the one matching current time
    current_period = None
    active_shift   = None
    for sh in ['FN', 'AF']:
        p = get_current_period(sh)
        if p:
            current_period = p
            active_shift   = sh
            break

    active_entry = None
    students     = []

    if current_period and active_shift:
        active_entry = db.execute("""
            SELECT tt.id, tt.period, tt.shift, tt.year, tt.section,
                   s.name as sub_name, s.id as subject_id
            FROM timetable tt JOIN subjects s ON tt.subject_id=s.id
            WHERE tt.teacher_id=? AND tt.day=? AND tt.period=? AND tt.shift=?
        """, (teacher_id, today_name, current_period, active_shift)).fetchone()

        if active_entry:
            # Check if already submitted
            already_done = db.execute(
                "SELECT id FROM attendance WHERE teacher_id=? AND date=? AND period=? AND subject_id=?",
                (teacher_id, today_date, current_period, active_entry['subject_id'])
            ).fetchone()
            if already_done:
                active_entry = None  # Marked as done
                flash('✅ Attendance for this period already submitted.', 'info')
            else:
                students = db.execute(
                    "SELECT * FROM students WHERE year=? AND section=? AND shift=? ORDER BY reg_no",
                    (active_entry['year'], active_entry['section'], active_entry['shift'])
                ).fetchall()

    # Recent history for this teacher
    history = db.execute("""
        SELECT a.date, a.period, s.name as sub_name, COUNT(a.id) as count
        FROM attendance a JOIN subjects s ON a.subject_id=s.id
        WHERE a.teacher_id=?
        GROUP BY a.date, a.period, a.subject_id
        ORDER BY a.date DESC, a.period LIMIT 10
    """, (teacher_id,)).fetchall()

    db.close()
    period_label = get_period_label(active_shift, current_period) if (current_period and active_shift) else None
    return render_template('teacher_dashboard.html',
                           day=today_name,
                           current_period=current_period,
                           active_shift=active_shift,
                           period_label=period_label,
                           active=active_entry,
                           students=students,
                           history=history)


@app.route('/submit_attendance', methods=['POST'])
@login_required('teacher')
def submit_attendance():
    teacher_id  = session['user_id']
    subject_id  = int(request.form['subject_id'])
    period      = int(request.form['period'])
    today_date  = date.today().isoformat()
    db          = get_db()
    reg_nos     = request.form.getlist('reg_nos')
    inserted    = 0
    for reg_no in reg_nos:
        status = request.form.get(f'status_{reg_no}', 'A')
        db.execute(
            "INSERT INTO attendance (student_id,subject_id,teacher_id,date,period,status) VALUES (?,?,?,?,?,?)",
            (reg_no, subject_id, teacher_id, today_date, period, status)
        )
        inserted += 1
    db.commit()
    db.close()
    flash(f'✅ Attendance saved for {inserted} students.', 'success')
    return redirect(url_for('teacher_dashboard'))


# ─────────────────────────────────────────────────────────────────
# STUDENT — DASHBOARD
# ─────────────────────────────────────────────────────────────────
@app.route('/student_dashboard')
@login_required('student')
def student_dashboard():
    db       = get_db()
    reg_no   = session['username']

    student  = db.execute("SELECT * FROM students WHERE reg_no=?", (reg_no,)).fetchone()
    if not student:
        flash('Student record not found.', 'danger')
        return redirect('/')

    # Subject-wise attendance
    subjects = db.execute(
        "SELECT * FROM subjects WHERE year=? ORDER BY semester,name",
        (student['year'],)
    ).fetchall()

    sub_data = []
    total_p  = 0
    total_c  = 0
    at_risk_subjects = []
    
    for sub in subjects:
        records = db.execute(
            "SELECT status FROM attendance WHERE student_id=? AND subject_id=?",
            (reg_no, sub['id'])
        ).fetchall()
        countable = [r for r in records if r['status'] in ('P', 'A')]
        t = len(countable)
        p = sum(1 for r in countable if r['status'] == 'P')
        od = sum(1 for r in records if r['status'] == 'OD')
        lv = sum(1 for r in records if r['status'] == 'L')
        pct = calc_percentage(p, t)
        total_p += p; total_c += t
        
        needs_classes = classes_needed_for_75(p, t)
        if pct < 75 and t > 0:
            at_risk_subjects.append({
                'name': sub['name'],
                'current_pct': pct,
                'needs': needs_classes
            })
            
        sub_data.append({
            'name': sub['name'],
            'semester': sub['semester'],
            'present': p, 'total': t, 'od': od, 'lv': lv, 'pct': pct,
            'color': status_color(pct)
        })

    overall_pct = calc_percentage(total_p, total_c)
    overall_needs = classes_needed_for_75(total_p, total_c)

    # Recent attendance history
    history = db.execute("""
        SELECT a.date, a.period, sub.name as sub_name, a.status
        FROM attendance a JOIN subjects sub ON a.subject_id=sub.id
        WHERE a.student_id=?
        ORDER BY a.date DESC, a.period DESC LIMIT 30
    """, (reg_no,)).fetchall()

    db.close()
    return render_template('student_dashboard.html',
                           student=student,
                           sub_data=sub_data,
                           overall_pct=overall_pct,
                           total_p=total_p,
                           total_c=total_c,
                           history=history,
                           at_risk_subjects=at_risk_subjects,
                           overall_needs=overall_needs,
                           color=status_color(overall_pct))


@app.route('/student/export_pdf')
@login_required('student')
def student_export_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib import colors
    except ImportError:
        flash('reportlab missing. pip install reportlab', 'danger')
        return redirect(url_for('student_dashboard'))
        
    db     = get_db()
    reg_no = session['username']
    student = db.execute("SELECT * FROM students WHERE reg_no=?", (reg_no,)).fetchone()
    if not student:
        db.close()
        flash('Student record not found.', 'danger')
        return redirect(url_for('student_dashboard'))

    subjects = db.execute(
        "SELECT * FROM subjects WHERE year=? ORDER BY semester,name", (student['year'],)
    ).fetchall()

    sub_data = []
    total_p = total_c = 0
    for sub in subjects:
        records = db.execute(
            "SELECT status FROM attendance WHERE student_id=? AND subject_id=?",
            (reg_no, sub['id'])
        ).fetchall()
        countable = [r for r in records if r['status'] in ('P', 'A')]
        t = len(countable)
        p = sum(1 for r in countable if r['status'] == 'P')
        od = sum(1 for r in records if r['status'] == 'OD')
        lv = sum(1 for r in records if r['status'] == 'L')
        pct = calc_percentage(p, t)
        total_p += p; total_c += t
        sub_data.append((sub['name'], sub['semester'], p, t, od, lv, pct))

    overall_pct = calc_percentage(total_p, total_c)
    db.close()

    buf  = io.BytesIO()
    doc  = SimpleDocTemplate(buf, pagesize=A4,
                             topMargin=2*colors.cm, bottomMargin=2*colors.cm,
                             leftMargin=2*colors.cm, rightMargin=2*colors.cm)
    styles = getSampleStyleSheet()
    story  = []

    # Title
    title_style = ParagraphStyle('title', parent=styles['Title'],
                                  fontSize=16, spaceAfter=4)
    story.append(Paragraph('BCA Attendance System', title_style))
    story.append(Paragraph('Personal Attendance Report', styles['Normal']))
    story.append(Spacer(1, 14))

    # Student info block
    info_data = [
        ['Name', student['name'], 'Reg No', student['reg_no']],
        ['Year', str(student['year']), 'Section', student['section']],
        ['Semester', str(student['semester']), 'Shift', student['shift']],
        ['Overall %', f'{overall_pct}%']
    ]
    info_tbl = Table(info_data, colWidths=[60, 150, 60, 150])
    info_tbl.setStyle(TableStyle([
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME', (2,0), (2,-1), 'Helvetica-Bold'),
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#f8fafc')),
        ('BOX', (0,0), (-1,-1), 0.5, colors.lightgrey),
        ('INNERGRID', (0,0), (-1,-1), 0.25, colors.lightgrey),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(info_tbl)
    story.append(Spacer(1, 20))

    # Subject-wise table
    table_data = [['Subject', 'Sem', 'Pres', 'Tot', 'OD', 'Leave', '%']]
    for row in sub_data:
        table_data.append([row[0], row[1], row[2], row[3], row[4], row[5], f'{row[6]}%'])

    tbl = Table(table_data, colWidths=[200, 40, 40, 40, 40, 40, 50])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 9),
        ('FONTSIZE',   (0,1), (-1,-1), 8.5),
        ('ROWBACKGROUNDS', (0,1), (-1,-2), [colors.white, colors.HexColor('#f8fafc')]),
        ('BACKGROUND', (0,-1), (-1,-1), colors.HexColor('#e0e7ff')),
        ('FONTNAME',   (0,-1), (-1,-1), 'Helvetica-Bold'),
        ('GRID',       (0,0), (-1,-1), 0.3, colors.lightgrey),
        ('ALIGN',      (1,0), (-1,-1), 'CENTER'),
    ]))
    # Colour the % columns
    for i, (_, _, p, t, od, lv, pct) in enumerate(sub_data, 1):
        if pct >= 75:
            tbl.setStyle(TableStyle([('TEXTCOLOR', (6,i),(6,i), colors.HexColor('#059669'))]))
        elif pct >= 50:
            tbl.setStyle(TableStyle([('TEXTCOLOR', (6,i),(6,i), colors.HexColor('#b45309'))]))
        else:
            tbl.setStyle(TableStyle([('TEXTCOLOR', (6,i),(6,i), colors.HexColor('#dc2626'))]))

    story.append(tbl)
    story.append(Spacer(1, 14))
    story.append(Paragraph(
        '<i>Note: OD (On Duty) and Leave records are not counted in attendance percentage.</i>',
        styles['Normal']
    ))

    doc.build(story)
    buf.seek(0)
    fname = f'Attendance_{reg_no}_{date.today().isoformat()}.pdf'
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype='application/pdf')


# ─────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    if not os.path.exists(DB_PATH):
        print("⚠ Database not found. Run init_db.py first.")
    init_admin()
    app.run(debug=True, port=5000)